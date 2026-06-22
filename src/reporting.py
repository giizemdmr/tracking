"""
reporting.py - Raporlama ve Analiz Modülü
Bu modül, araçların bölgeler arası geçişlerini takip eder ve Excel raporu oluşturur.
"""

import cv2
import numpy as np
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional
import os
import math


def extract_video_start_time(video_path: str) -> Optional[datetime]:
    """
    ffprobe ile video dosyasının oluşturulma zamanını çeker.
    Önce metadata'dan, bulamazsa dosya adından parse etmeye çalışır.
    
    Returns:
        datetime nesnesi veya None
    """
    # 1. ffprobe ile metadata'dan çek
    try:
        result = subprocess.run(
            ['ffprobe', '-v', 'quiet', '-print_format', 'json', '-show_format', video_path],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            data = json.loads(result.stdout)
            tags = data.get('format', {}).get('tags', {})
            creation_time = tags.get('creation_time') or tags.get('com.apple.quicktime.creationdate')
            if creation_time:
                # ISO format: "2026-06-15T07:35:51.000000Z"
                ct = creation_time.replace('Z', '+00:00')
                dt = datetime.fromisoformat(ct)
                # UTC'den yerel saate çevir (Türkiye UTC+3)
                dt = dt + timedelta(hours=3)
                print(f"[OK] Video başlangıç zamanı (metadata): {dt.strftime('%Y-%m-%d %H:%M:%S')}")
                return dt
    except Exception as e:
        print(f"[WARNING] ffprobe metadata okunamadı: {e}")
    
    # 2. Dosya adından parse et (örnek: 2026_0615_073551_006.MP4)
    try:
        basename = os.path.splitext(os.path.basename(video_path))[0]
        parts = basename.split('_')
        if len(parts) >= 3 and len(parts[0]) == 4 and len(parts[1]) == 4 and len(parts[2]) == 6:
            year = int(parts[0])
            month = int(parts[1][:2])
            day = int(parts[1][2:])
            hour = int(parts[2][:2])
            minute = int(parts[2][2:4])
            second = int(parts[2][4:])
            dt = datetime(year, month, day, hour, minute, second)
            print(f"[OK] Video başlangıç zamanı (dosya adı): {dt.strftime('%Y-%m-%d %H:%M:%S')}")
            return dt
    except Exception as e:
        print(f"[WARNING] Dosya adından zaman parse edilemedi: {e}")
    
    print("[WARNING] Video başlangıç zamanı belirlenemedi. Video-göreceli zaman kullanılacak.")
    return None

class VehicleRoute:
    """Bir aracın izlediği rota (Gate-to-Gate) bilgilerini tutar."""
    def __init__(self, object_id: int, config=None):
        self.object_id = object_id
        self.config = config
        self.vehicle_type = "Bilinmiyor"
        
        self.entry_gate = None  # Mühürlenmiş Origin Hattı
        self.exit_gate = None   # Son Geçilen Destination Hattı
        self.is_completed = False
        
        # Frame-tabanlı zaman damgası
        self.entry_frame = None  # İlk kapı teması frame numarası
        self.exit_frame = None   # Son kapı teması frame numarası
        
        # O-D Mantığı Veri Yapıları
        self.gate_counts = {}   # {gate_name: frame_count}
        self.gate_last_seen = {} # {gate_name: last_frame_id}
        self.route = []         # Görsel rapor için kronolojik geçişler
        
    def add_gate_contact(self, gate_name: str, frame_id: int) -> bool:
        """
        Kati O-D Kurallari:
        1. Origin (Giriş) hattı ilk temasla mühürlenir.
        2. Her temas için frame sayılır ve kronolojik rota tutulur.
        """
        if gate_name is None:
            return False

        # --- 1. ORIGIN KİLİDİ ---
        if self.entry_gate is None:
            self.entry_gate = gate_name
            self.entry_frame = frame_id

        # --- 2. VERİ TOPLAMA ---
        self.gate_counts[gate_name] = self.gate_counts.get(gate_name, 0) + 1
        self.gate_last_seen[gate_name] = frame_id
        self.exit_frame = frame_id  # Her temasta güncelle, son temas = çıkış
        
        # Görsel rota dizisi için (ardışık tekrarları önle)
        if not self.route or self.route[-1] != gate_name:
            self.route.append(gate_name)
        
        return True

    def finalize(self):
        """Alternatif çıkış belirleme ve rotayı tamamlama."""
        if not self.entry_gate:
            return
            
        # Eğer birden fazla hat ile temas varsa, girişten farklı en son hattı 'Çıkış' yap.
        if len(self.gate_counts) > 1:
            for gate in reversed(self.route):
                if gate != self.entry_gate:
                    self.exit_gate = gate
                    break
        
        self.is_completed = True

    def get_report(self) -> str:
        """Rota özetini metin olarak döndürür."""
        origin = self.entry_gate or "Bilinmiyor"
        dest = self.exit_gate or "Bilinmiyor"
        return f"ID {self.object_id} [{self.vehicle_type}]: {origin} -> {dest} (Rota: {' -> '.join(self.route)})"

    def get_route_string(self) -> str:
        """Tam rotayı string olarak döndürür."""
        return " -> ".join(self.route)

class VehicleTracker:
    """Tüm araçların ömür döngüsünü ve hat geçişlerini yönetir."""
    def __init__(self, config):
        self.config = config
        self.active_vehicles: Dict[int, VehicleRoute] = {}
        self.type_stats: Dict[int, Dict[str, float]] = {} # ID -> {type: total_conf}
        self.last_seen_frame: Dict[int, int] = {} # ID -> frame_id
        self.total_vehicles_seen = 0

    def update_liveness(self, object_id: int, frame_id: int) -> None:
        """Her frame'de aracin hala ekranda oldugunu kaydeder."""
        self.last_seen_frame[object_id] = frame_id

    def update_gate_contact(self, object_id: int, gate_name: str, frame_id: int) -> None:
        """Aracın bir hattan geçtiğini veya hattı tetiklediğini kaydeder."""
        if object_id not in self.active_vehicles:
            self.active_vehicles[object_id] = VehicleRoute(object_id, self.config)
            self.total_vehicles_seen += 1
            
        self.active_vehicles[object_id].add_gate_contact(gate_name, frame_id)

    def update_type_score(self, object_id: int, vehicle_type: str, confidence: float, bbox: List[int] = None) -> None:
        """
        Scale-Aware (Alan Agirlikli) Sınıf Oylama Mekanizması.
        
        Aracın güven skoru (confidence) ile bounding box alanını harmanlayarak bir ağırlık hesaplar.
        Büyük alanlı (kameraya yakın) ve yüksek güvenli tespitler oylamada daha etkili olur.
        """
        if object_id not in self.type_stats:
            self.type_stats[object_id] = {'high_conf': {}, 'all': {}}
            
        # Merkezi konfigürasyondan eşik değerini çek (0.75 varsayılan)
        vote_threshold = getattr(self.config, 'class_vote_threshold', 0.75)
            
        # 1. Alan Ağırlıklı (Scale-Aware) Ağırlık Hesaplama
        weight = confidence
        if bbox is not None and len(bbox) >= 4:
            x1, y1, x2, y2 = bbox
            area = abs(x2 - x1) * abs(y2 - y1)
            # Pikselleri normalize et ve karekök sönümlendirmesi uygula (Scale-Aware Math)
            norm_area = area / 10000.0
            weight = confidence * math.sqrt(norm_area)

        # 2. Eşik Değer Kontrolü (Kaliteli Tespit Filtresi)
        if confidence >= vote_threshold:
            self.type_stats[object_id]['high_conf'][vehicle_type] = self.type_stats[object_id]['high_conf'].get(vehicle_type, 0.0) + weight
        
        # 3. Fallback Listesi (Basit kümülatif toplama devam eder)
        self.type_stats[object_id]['all'][vehicle_type] = self.type_stats[object_id]['all'].get(vehicle_type, 0.0) + confidence

    def get_best_type(self, object_id: int) -> str:
        """Önce eşik üzeri kesin kayıtlara bakar, en yüksek toplam skoru alana galibiyeti verir."""
        if object_id in self.type_stats:
            high_scores = self.type_stats[object_id]['high_conf']
            fallback_scores = self.type_stats[object_id]['all']
            
            # Eğer eşik üzeri tek bir tespiti dahi varsa, o gruba ait skor liderini sec!
            if high_scores:
                return max(high_scores, key=high_scores.get)
            # Arac 100 frame boyunca gozuktu ama asla 0.70 skoruna cikamadiysa (sis, karanlik vs)
            elif fallback_scores:
                return max(fallback_scores, key=fallback_scores.get)
                
        return "Bilinmiyor"

    def get_completed_routes(self, current_frame: int, stale_buffer: int = 200) -> List[VehicleRoute]:
        """Ekrandan tamamen kaybolan (stale olan) araçların rotalarını döndürür ve temizler."""
        completed = []
        ids_to_remove = []
        
        for tid, route in self.active_vehicles.items():
            last_seen = self.last_seen_frame.get(tid, 0)
            if current_frame - last_seen > stale_buffer:
                # Araç tamamen ekrandan çıktı (stale oldu)
                route.finalize()
                route.vehicle_type = self.get_best_type(tid)
                completed.append(route)
                ids_to_remove.append(tid)
                
        for tid in ids_to_remove:
            if tid in self.active_vehicles: del self.active_vehicles[tid]
            if tid in self.type_stats: del self.type_stats[tid]
            if tid in self.last_seen_frame: del self.last_seen_frame[tid]
            
        return completed


class ReportGenerator:
    """
    Trafik analiz raporlarını Excel formatında oluşturur.
    
    Araç geçişlerini kaydeder ve Excel dosyasına yazar.
    D:\\report_generator.py mantığı birebir uygulanmıştır.
    
    Attributes:
        config: Uygulama konfigürasyonu
        records: Rapor kayıtları listesi
    """
    
    def __init__(self, config):
        """
        ReportGenerator başlatıcısı.
        
        Args:
            config: Uygulama konfigürasyonu
        """
        self.config = config
        self.records: List[List] = []
        
        # Video zaman damgası için metadata
        self.fps: float = 30.0
        self.vid_stride: int = 1
        self.video_start_time: Optional[datetime] = None
    
    def set_video_metadata(self, fps: float, vid_stride: int, video_path: str) -> None:
        """
        Video metadata'sını ayarlar. ffprobe ile başlangıç zamanını çeker.
        
        Args:
            fps: Video FPS değeri
            vid_stride: Kare atlama adımı
            video_path: Video dosya yolu (ffprobe için)
        """
        self.fps = fps
        self.vid_stride = vid_stride
        self.video_start_time = extract_video_start_time(video_path)
    
    def _frame_to_timestamp(self, frame_id: Optional[int]) -> str:
        """
        İşlenmiş frame numarasını gerçek zaman damgasına çevirir.
        
        Hesaplama: video_start_time + (frame_id * vid_stride / fps)
        """
        if frame_id is None:
            return ""
        
        video_seconds = (frame_id * self.vid_stride) / self.fps
        
        if self.video_start_time:
            actual_time = self.video_start_time + timedelta(seconds=video_seconds)
            return actual_time.strftime("%H:%M:%S")
        else:
            # Fallback: video-göreceli zaman (HH:MM:SS)
            hours = int(video_seconds // 3600)
            minutes = int((video_seconds % 3600) // 60)
            seconds = int(video_seconds % 60)
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    def add_route(self, route: 'VehicleRoute') -> None:
        """
        Tamamlanmış bir rota ekler.
        Bilinmeyen çıkış kapısı olan rotalar (tek kapı teması) Excel'e kaydedilmez.
        
        Args:
            route: VehicleRoute nesnesi
        """
        # Çıkış kapısı belirsiz olan (tek kapıya temas edip kaybolan) araçları atla
        if route.exit_gate is None:
            return
        
        entry_time = self._frame_to_timestamp(route.entry_frame)
        route_string = route.get_route_string()
        
        record = [
            entry_time,
            route.object_id,
            route.vehicle_type,
            route.entry_gate,
            route.exit_gate,
            route_string
        ]
        
        self.records.append(record)
    
    def add_record(
        self,
        object_id: int,
        vehicle_type: str,
        entry_gate: str,
        exit_gate: str,
        full_route: List[str]
    ) -> None:
        """
        Manuel olarak kayıt ekler.
        
        Args:
            object_id: Araç ID'si
            vehicle_type: Araç tipi
            entry_gate: Giriş hattı
            exit_gate: Çıkış hattı
            full_route: Tam rota listesi
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        # Giris ve Cikis hatlarini al
        entry_gate = entry_gate or "Unknown"
        exit_gate = exit_gate or "Unknown"
        route_string = " -> ".join(full_route)

        record = [
            timestamp,
            object_id,
            vehicle_type,
            entry_gate,
            exit_gate,
            route_string
        ]
        
        self.records.append(record)

    def get_record_count(self) -> int:
        """Kayıt sayısını döndürür."""
        return len(self.records)
    
    def save_final_report(self, tracker: 'VehicleTracker') -> bool:
        """
        Video sonunda aktif kalan tüm araçların rotalarını 'boş tespit listesi' ile zorla (-flush) 
        kapatarak rapora ekler ve projeyi kaydeder.
        """
        print("\n[INFO] Video sonlandi. Yarim kalan rotalar kapatiliyor...")
        
        # Tracker'a cok yuksek bir frame yollayip butun aktif araclari (stale) hale getirip bitirt.
        flushed_routes = tracker.get_completed_routes(current_frame=9999999, stale_buffer=0)
        for route in flushed_routes:
            self.add_route(route)
            print(f"[FLUSH TAMAMLANDI] {route.get_report()}")
            
        return self.save()

    def save(self) -> bool:
        """
        Raporu Excel dosyasına kaydeder.
        
        Returns:
            True eğer başarılı, False eğer hata oluşursa
        """
        print("\n[INFO] Excel raporu olusturuluyor...")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trafik Raporu"
            
            # Başlıklar
            headers = [
                "Zaman",
                "ID",
                "Tür",
                "Giriş Kapısı",
                "Çıkış Kapısı",
                "Tam Rota"
            ]
            ws.append(headers)
            
            # Başlık stili
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_text = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.font = header_text
                cell.fill = header_fill
            
            # Veri satırları
            for record in self.records:
                ws.append(record)
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 20
            
            # Kaydet
            wb.save(self.config.excel_filename)
            print(f"[OK] {len(self.records)} kayit yazildi: {self.config.excel_filename}")
            
            return True
            
        except PermissionError:
            print(f"[ERROR] Dosya erisim hatasi: {self.config.excel_filename}")
            print("   Dosya baska bir program tarafindan acik olabilir.")
            return False
            
        except Exception as e:
            print(f"[ERROR] Excel hatasi: {e}")
            return False
    
    def print_statistics(self, frame_count: int, vehicle_count: int = 0) -> None:
        """
        İstatistikleri konsola yazdırır.
        
        Args:
            frame_count: Toplam frame sayısı
            vehicle_count: Tespit edilen toplam araç sayısı
        """
        print("\n" + "=" * 60)
        print("[STATS] ISTATISTIKLER")
        print("=" * 60)
        print(f"Toplam Frame: {frame_count}")
        print(f"Tespit Edilen Araç: {vehicle_count}")
        print(f"Raporlanan Geçiş: {len(self.records)}")
        print("=" * 60)
    
    def clear(self) -> None:
        """Tüm kayıtları temizler."""
        self.records.clear()
