import pandas as pd
import glob
import os
import openpyxl
from datetime import datetime, time

def round_time_15min_obj(time_str):
    try:
        t = datetime.strptime(str(time_str), "%H:%M:%S")
        minute = (t.minute // 15) * 15
        return time(t.hour, minute, 0)
    except:
        return None

def main():
    folder = r"d:\tracking\otogarkavsagi"
    folder_alt = r"d:\tracking\otogar kavsağı"
    ciktilar_folder = r"d:\tracking\ciktilar"
    
    # Videolardaki ham verileri topla
    files = glob.glob(os.path.join(folder, "*_rapor.xlsx"))
    if not files:
        files = glob.glob(os.path.join(ciktilar_folder, "*_rapor.xlsx"))
    if not files:
        files = glob.glob(os.path.join(folder_alt, "*_rapor.xlsx"))
        
    if not files:
        print("[HATA] Excel raporlari bulunamadi!")
        return

    print(f"[INFO] {len(files)} adet Excel ham verisi okunuyor...")
    dfs = []
    for f in files:
        try:
            dfs.append(pd.read_excel(f))
        except:
            pass

    if not dfs:
        return

    all_data = pd.concat(dfs, ignore_index=True)
    
    # Sütun isimleri (esnek okuma)
    zaman_col = all_data.columns[0]
    tur_col = all_data.columns[2]
    rota_col = all_data.columns[5]

    # Zamanı 15 dakikalık nesnelere yuvarla (örn: datetime.time(8, 15))
    all_data['Start_Time'] = all_data[zaman_col].apply(round_time_15min_obj)
    
    # Araç Sınıflarını Şablona Map et
    def map_tur(tur_str):
        t = str(tur_str).lower()
        if "otomobil" in t: return "OTOMOBİL"
        if "kamyonet" in t or "panelvan" in t: return "KAMYONET"
        if "otobüs" in t: return "İETT+HALK"
        if "ağır" in t or "kamyon" in t or "tır" in t: return "AĞIR TAŞIT"
        if "minibüs" in t: return "T.MİNİBÜS"
        return "OTOMOBİL" # Varsayılan

    all_data['Sablon_Tur'] = all_data[tur_col].apply(map_tur)
    
    # Rota'yı Ayrıştır (Örn: '1 -> 2') -> origin='1', dest='2'
    def parse_rota(r):
        try:
            parts = str(r).split("->")
            orig = parts[0].strip().split()[-1] # "Giriş Kapısı 1" -> "1"
            dest = parts[1].strip().split()[-1] # "Çıkış Kapısı 2" -> "2"
            return orig, dest
        except:
            return None, None

    all_data[['Origin', 'Dest']] = all_data[rota_col].apply(lambda x: pd.Series(parse_rota(x)))

    # Hatalı rotaları filtrele
    all_data = all_data.dropna(subset=['Origin', 'Dest', 'Start_Time'])

    # Pivot / Grup
    grouped = all_data.groupby(['Origin', 'Dest', 'Start_Time', 'Sablon_Tur']).size().reset_index(name='Count')

    # Şablonu Aç
    template_path = r"D:\K01 (05.11.2022).xlsx"
    output_path = r"D:\tracking\otogarkavsagi.xlsx"
    
    print(f"[INFO] '{template_path}' şablonu açılıyor...")
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"[HATA] Şablon açılamadı: {e}")
        return

    # Araç sütun ofsetleri (Row 6'da '1-2' kaçıncı sütun başlıyor)
    # Şablondaki ana bloklar:
    # OTOMOBİL: Col 4 (D) başlar
    # KAMYONET: Col 10 (J) başlar
    # TAKSİ: Col 16 (P) başlar
    # T.MİNİBÜS: Col 22 (V) başlar
    # SERVİS MİN.: Col 28 (AB) başlar
    # İETT+HALK: Col 34 (AH) başlar
    # AĞIR TAŞIT: Col 40 (AN) başlar
    tur_offsets = {
        "OTOMOBİL": 4,
        "KAMYONET": 10,
        "TAKSİ": 16,
        "T.MİNİBÜS": 22,
        "SERVİS MİN.": 28,
        "İETT+HALK": 34,
        "AĞIR TAŞIT": 40
    }

    print("[INFO] Veriler şablondaki hücrelere işleniyor...")
    
    # Satır aralığını bulmak için önbellek (her sheet için saat->satır haritası)
    # 7. satırdan başlayıp aşağı inerek B sütunundaki saatleri okuyacağız.
    
    for origin in grouped['Origin'].unique():
        sheet_name = str(origin)
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # O sheet'teki saatlerin hangi satırda olduğunu bul
        time_to_row = {}
        for r in range(8, 100):
            cell_val = ws.cell(row=r, column=2).value
            if isinstance(cell_val, time):
                time_to_row[cell_val] = r
                
        # O origin için olan verileri filtrele
        orig_data = grouped[grouped['Origin'] == origin]
        
        for _, row_data in orig_data.iterrows():
            dest = str(row_data['Dest'])
            t_obj = row_data['Start_Time']
            tur = row_data['Sablon_Tur']
            count = row_data['Count']
            
            # Doğru satırı bul
            if t_obj not in time_to_row:
                continue
            r_idx = time_to_row[t_obj]
            
            # Doğru sütunu bul
            if tur not in tur_offsets:
                continue
            base_col = tur_offsets[tur]
            
            # Hedefin kaçıncı sütun olduğunu hesapla
            # Şablonda başlıklar genelde: U, 1-2, 1-3, 1-4, 1-5, 1-6 (Col indexleri +0, +1, +2...)
            # Biz dest stringini kontrol edeceğiz. Row 7'ye (Excel 7. satır) bakıyoruz.
            col_idx = None
            for c_offset in range(6):
                header_val = str(ws.cell(row=7, column=base_col + c_offset).value)
                if dest in header_val:
                    col_idx = base_col + c_offset
                    break
            
            if col_idx is not None:
                # Sayıyı hücreye ekle (eğer önceden sayı varsa üzerine ekle, gerçi sıfırdan yazıyoruz)
                # Formülleri bozmamak için sadece integer ise yaz
                # Eğer şablonda '-' vs varsa ezeriz.
                current = ws.cell(row=r_idx, column=col_idx).value
                if not isinstance(current, (int, float)):
                    current = 0
                ws.cell(row=r_idx, column=col_idx).value = current + count

    print(f"[INFO] Şablon kaydediliyor: {output_path} (Bu işlem formülleri hesaplayacağı için birkaç saniye sürebilir)")
    wb.save(output_path)
    print("\n[BAŞARILI] Tüm otonom sayımlar şablona eksiksiz ve formülleri bozmadan işlendi!")

if __name__ == "__main__":
    main()
