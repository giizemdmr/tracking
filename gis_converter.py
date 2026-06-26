import os
import sys
import glob
import re
import openpyxl
import copy
import pandas as pd
from datetime import datetime, timedelta, time

# Varsayılan yollar (Argüman girilmezse kullanılacak)
TEMPLATE_PATH = r"D:\tracking\mersankavsagi.xlsx"
DEFAULT_INPUT_DIR = r"D:\sayım sonuçları-uni\sayım sonuçları"
DEFAULT_OUTPUT_PATH = r"D:\sayım sonuçları-uni\unikavsagi.xlsx"

# Komut satırı argümanlarını kontrol et
if len(sys.argv) > 2:
    INPUT_DIR = sys.argv[1]
    OUTPUT_PATH = sys.argv[2]
else:
    INPUT_DIR = DEFAULT_INPUT_DIR
    OUTPUT_PATH = DEFAULT_OUTPUT_PATH

category_mapping = {
    'otomobil': 'OTOMOBIL',
    'kamyonet': 'KAMYONET',
    'panelvan': 'PANELVAN',
    'minibus': 'MINIBUS',
    'otobus': 'OTOBUS',
    'agir_tasit': 'AGIR_TASIT'
}

exclude_turs = ["bisiklet", "motosiklet", "yaya"]

def get_gate_num(gate_name):
    if pd.isna(gate_name):
        return None
    match = re.search(r'\d+', str(gate_name))
    if match:
        return int(match.group())
    return None

def get_file_time(filename):
    basename = os.path.basename(filename)
    digits = re.sub(r'\D', '', basename)[:14]
    return datetime.strptime(digits, "%Y%m%d%H%M%S")

def get_excel_duration(df):
    if 'Zaman' not in df.columns or len(df) == 0:
        return timedelta(minutes=15)
    try:
        times = []
        for val in df['Zaman']:
            if pd.isna(val):
                continue
            if isinstance(val, time):
                times.append(datetime.combine(datetime.today(), val))
            elif isinstance(val, datetime):
                times.append(val)
            else:
                for fmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        times.append(datetime.strptime(str(val).strip(), fmt))
                        break
                    except ValueError:
                        continue
        if not times:
            return timedelta(minutes=15)
        min_t = min(times)
        max_t = max(times)
        return max_t - min_t
    except Exception:
        return timedelta(minutes=15)

def round_time_to_15m(dt):
    minute = dt.minute
    second = dt.second
    total_minutes = minute + second / 60.0
    rounded_minute = int(15 * round(total_minutes / 15.0))
    
    if rounded_minute == 60:
        dt = dt + timedelta(hours=1)
        dt = dt.replace(minute=0, second=0, microsecond=0)
    else:
        dt = dt.replace(minute=rounded_minute, second=0, microsecond=0)
    return dt

def adjust_merged_column_a(ws, start_row, n_sabah, n_ogle, end_row=32):
    # Column A içindeki tüm birleştirilmiş hücreleri çöz
    ranges_to_remove = []
    for rng in list(ws.merged_cells.ranges):
        if rng.bounds[0] == 1 and rng.bounds[2] == 1: # Sadece A Sütunu (1)
            ranges_to_remove.append(rng)
    for rng in ranges_to_remove:
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass
            
    # Mevcut A sütunu değerlerini temizle
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=1).value = None

    # SABAH birleştir
    r_sabah_start = start_row
    r_sabah_end = start_row + n_sabah - 1
    if r_sabah_end >= r_sabah_start:
        ws.cell(row=r_sabah_start, column=1).value = "SABAH"
        if r_sabah_end > r_sabah_start:
            ws.merge_cells(start_row=r_sabah_start, start_column=1, end_row=r_sabah_end, end_column=1)

    # ÖĞLE birleştir
    r_ogle_start = r_sabah_end + 1
    r_ogle_end = r_sabah_end + n_ogle
    if r_ogle_end >= r_ogle_start:
        ws.cell(row=r_ogle_start, column=1).value = "ÖĞLE"
        if r_ogle_end > r_ogle_start:
            ws.merge_cells(start_row=r_ogle_start, start_column=1, end_row=r_ogle_end, end_column=1)

    # AKŞAM birleştir
    r_aksam_start = r_ogle_end + 1
    r_aksam_end = end_row
    if r_aksam_end >= r_aksam_start:
        ws.cell(row=r_aksam_start, column=1).value = "AKŞAM"
        if r_aksam_end > r_aksam_start:
            ws.merge_cells(start_row=r_aksam_start, start_column=1, end_row=r_aksam_end, end_column=1)

def adjust_row_1_merge(ws, base_col, new_toplam_col):
    # Row 1'de base_col ile çakışan merged range varsa unmerge et
    ranges_to_remove = []
    for rng in list(ws.merged_cells.ranges):
        if rng.bounds[1] == 1 and rng.bounds[3] == 1: # Row 1
            if rng.bounds[0] <= base_col <= rng.bounds[2]:
                ranges_to_remove.append(rng)
    for rng in ranges_to_remove:
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass
    # Yeni genişliği merge et
    ws.merge_cells(start_row=1, start_column=base_col, end_row=1, end_column=new_toplam_col)

def copy_row_style(ws, src_row, dst_row):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        cell_src = ws.cell(row=src_row, column=col)
        cell_dst = ws.cell(row=dst_row, column=col)
        if cell_src.has_style:
            cell_dst.font = copy.copy(cell_src.font)
            cell_dst.fill = copy.copy(cell_src.fill)
            cell_dst.border = copy.copy(cell_src.border)
            cell_dst.alignment = copy.copy(cell_src.alignment)
            cell_dst.number_format = cell_src.number_format

def main():
    print("="*60)
    print(" KAVSAK REPORT CONVERTER: Excel Birleştirme (Mersan Formatı)")
    print(f" Girdi Klasörü: {INPUT_DIR}")
    print(f" Çıktı Dosyası: {OUTPUT_PATH}")
    print("="*60)

    # 1. Rapor dosyalarını yükle
    files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    if not files:
        print(f"[ERROR] '{INPUT_DIR}' klasörü içinde excel dosyası bulunamadı!")
        return

    # Dosyaları isimlerindeki zaman damgasına göre sırala
    files.sort(key=get_file_time)
    print(f"[INFO] '{INPUT_DIR}' klasöründen {len(files)} adet rapor kronolojik olarak okundu.")

    # Verileri birleştirerek dinamik özellikleri analiz et (Kapı sayıları vb.)
    dfs = []
    for f in files:
        try:
            dfs.append(pd.read_excel(f))
        except Exception as e:
            print(f"[WARNING] {os.path.basename(f)} okunamadı: {e}")
    if not dfs:
        print("[ERROR] Hiçbir veri okunamadı!")
        return
    all_data = pd.concat(dfs, ignore_index=True)

    # Dinamik Giriş ve Çıkış Kapılarını bul
    entry_gates = sorted([int(g) for g in all_data['Giriş Kapısı'].apply(get_gate_num).dropna().unique()])
    entry_gates = [g for g in entry_gates if g >= 1]
    
    exit_gates = sorted([int(g) for g in all_data['Çıkış Kapısı'].apply(get_gate_num).dropna().unique()])
    exit_gates = [g for g in exit_gates if g >= 1]
    
    max_exit_gate = max(exit_gates) if exit_gates else 8
    if max_exit_gate < 8:
        max_exit_gate = 8

    print(f"[INFO] Dinamik Akım Kapıları Tespit Edildi: {entry_gates}")
    print(f"[INFO] Maksimum Çıkış Yönü Sayısı: {max_exit_gate}")

    # Her dosyanın zaman aralığını dinamik olarak hesapla
    file_slots = {}
    sabah_files = []
    ogle_files = []
    aksam_files = []

    print("[INFO] Excel dosyalarından süreler ve zaman aralıkları hesaplanıyor...")
    for f in files:
        dt = get_file_time(f)
        start_dt = round_time_to_15m(dt)
        end_dt = start_dt + timedelta(minutes=15)
        
        slot = f"{start_dt.strftime('%H:%M')} - {end_dt.strftime('%H:%M')}"
        file_slots[f] = slot
        
        # Başlangıç saatine göre grupla
        start_hour = start_dt.hour
        if start_hour < 12:
            sabah_files.append(f)
        elif start_hour < 15:
            ogle_files.append(f)
        else:
            aksam_files.append(f)

    print(f"[INFO] Gruplama: SABAH={len(sabah_files)}, ÖĞLE={len(ogle_files)}, AKŞAM={len(aksam_files)}")
    sorted_files = sabah_files + ogle_files + aksam_files
    total_files = len(sorted_files)
    last_row_akim = 2 + total_files

    # 2. Şablon Excel Dosyasını Yükle
    if not os.path.exists(TEMPLATE_PATH):
        print(f"[ERROR] Şablon dosya bulunamadı: {TEMPLATE_PATH}")
        return

    print(f"[INFO] Şablon dosya yükleniyor: {TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    # Akım sayfaları listesini dinamik olarak belirle
    akim_sheets = [f"{i}. AKIM" for i in entry_gates]

    # 3. Şablonda bulunmayan akım sayfalarını otomatik kopyala ve oluştur
    src_sheet = None
    for s_name in wb.sheetnames:
        if "AKIM" in s_name:
            src_sheet = wb[s_name]
            break

    if src_sheet is None:
        print("[ERROR] Şablonda kopyalanacak hiçbir AKIM sayfası bulunamadı!")
        return

    for sheet_name in akim_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"[INFO] '{sheet_name}' şablonda yok, '{src_sheet.title}' kopyalanarak oluşturuluyor...")
            ws_new = wb.copy_worksheet(src_sheet)
            ws_new.title = sheet_name
            
            # Yeni kopyadaki 2. satır yön etiketlerini güncelle (Örn: 1-1 -> 9-1)
            g_num = get_gate_num(sheet_name)
            src_g_num = get_gate_num(src_sheet.title)
            prefix_to_replace = f"{src_g_num}-"
            
            for c in range(3, 100):
                val = ws_new.cell(row=2, column=c).value
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.startswith(prefix_to_replace):
                        ws_new.cell(row=2, column=c).value = val_str.replace(prefix_to_replace, f"{g_num}-")

    # Kullanılmayan akım sayfalarını şablondan temizle
    for sheet_name in list(wb.sheetnames):
        if "AKIM" in sheet_name and sheet_name not in akim_sheets:
            print(f"[INFO] '{sheet_name}' aktif akımlar arasında değil, şablondan siliniyor...")
            wb.remove(wb[sheet_name])

    # Sayfaları sıralayalım: AKIM sayfaları sayısal sırada olsun, ÖZET sayfası en sonda olsun
    def sheet_sort_key(sheet):
        name = sheet.title
        if "AKIM" in name:
            num = get_gate_num(name)
            return (0, num if num is not None else 0)
        else:
            return (1, 0)
    wb._sheets.sort(key=sheet_sort_key)

    # 4. Gerekirse Çıkış Kapısı Sayısına Göre Sütun Ekle (8 kapıdan fazla ise)
    if max_exit_gate > 8:
        num_new = max_exit_gate - 8
        print(f"[INFO] Çıkış kapısı sayısı 8'den büyük olduğu için {num_new} yeni sütun eklenecek...")
        
        for sheet_name in akim_sheets:
            ws = wb[sheet_name]
            g_num = get_gate_num(sheet_name)
            
            # Şablonun ilk blok yapısını bul
            vehicle_blocks = []
            current_vehicle = None
            base_col = None
            for c in range(3, 100):
                v_val = ws.cell(row=1, column=c).value
                if v_val is not None:
                    current_vehicle = str(v_val).strip().upper()
                    base_col = c
                d_val = ws.cell(row=2, column=c).value
                if d_val is not None:
                    d_val_str = str(d_val).strip().upper()
                    if d_val_str == 'TOPLAM':
                        if current_vehicle and base_col:
                            vehicle_blocks.append((current_vehicle, base_col, c))
                            
            # Sütun kaymalarını önlemek için sağdan sola doğru ekle
            for vehicle_name, base_c, toplam_c in reversed(vehicle_blocks):
                # TOPLAM sütununun soluna yeni sütunlar ekle
                ws.insert_cols(toplam_c, num_new)
                
                # Biçimlendirmeleri (Style) kopyala (Hemen soldaki gate_8 sütunundan)
                src_col = toplam_c - 1
                for offset in range(num_new):
                    dest_col = toplam_c + offset
                    for r in range(1, 33):
                        src_cell = ws.cell(row=r, column=src_col)
                        dest_cell = ws.cell(row=r, column=dest_col)
                        if src_cell.has_style:
                            dest_cell.font = copy.copy(src_cell.font)
                            dest_cell.fill = copy.copy(src_cell.fill)
                            dest_cell.border = copy.copy(src_cell.border)
                            dest_cell.alignment = copy.copy(src_cell.alignment)
                            dest_cell.number_format = src_cell.number_format
                            
                    # 2. satır çıkış etiketini yaz
                    exit_gate_num = 8 + 1 + offset
                    ws.cell(row=2, column=dest_col).value = f"{g_num}-{exit_gate_num}"
                    
                # 1. satır birleşik araç kategorisi alanını genişlet
                new_toplam_c = toplam_c + num_new
                adjust_row_1_merge(ws, base_c, new_toplam_c)

    # 5. Şablonun sütun yapılarını dinamik olarak oku (Güncellenmiş halini)
    col_mappings_by_sheet = {}
    total_cols_by_sheet = {}
    
    for sheet_name in akim_sheets:
        ws = wb[sheet_name]
        col_mappings = {}
        total_cols = {}
        current_vehicle = None
        
        for c in range(3, 120):
            v_val = ws.cell(row=1, column=c).value
            if v_val is not None:
                current_vehicle = str(v_val).strip().upper()
                
            d_val = ws.cell(row=2, column=c).value
            if d_val is not None:
                d_val_str = str(d_val).strip().upper()
                if d_val_str == 'TOPLAM':
                    if current_vehicle:
                        total_cols[current_vehicle] = c
                elif '-' in d_val_str:
                    parts = d_val_str.split('-')
                    try:
                        exit_gate = int(parts[1].strip())
                        if current_vehicle:
                            col_mappings[c] = (current_vehicle, exit_gate)
                    except ValueError:
                        pass
        col_mappings_by_sheet[sheet_name] = col_mappings
        total_cols_by_sheet[sheet_name] = total_cols

    # 6. Akım sayfalarını ilklendir ve zaman kısımlarını yaz
    for sheet_name in akim_sheets:
        ws = wb[sheet_name]
        
        # Dinamik satır ekleme/silme işlemi
        if total_files < 30:
            ws.delete_rows(last_row_akim + 1, 32 - last_row_akim)
        elif total_files > 30:
            ws.insert_rows(32, last_row_akim - 32)
            for r in range(32, last_row_akim + 1):
                copy_row_style(ws, 31, r)
        
        # Sütun A birleştirilmiş alanlarını güncelle
        adjust_merged_column_a(ws, start_row=3, n_sabah=len(sabah_files), n_ogle=len(ogle_files), end_row=last_row_akim)
        
        # Zaman ve verileri sıfırla/ilklendir
        max_col_idx = max(col_mappings_by_sheet[sheet_name].keys()) if col_mappings_by_sheet[sheet_name] else 56
        max_tot_idx = max(total_cols_by_sheet[sheet_name].values()) if total_cols_by_sheet[sheet_name] else 56
        limit_col = max(max_col_idx, max_tot_idx)
        
        for r in range(3, last_row_akim + 1):
            file_idx = r - 3
            f = sorted_files[file_idx]
            ws.cell(row=r, column=2).value = file_slots[f]
            for c in range(3, limit_col + 1):
                ws.cell(row=r, column=c).value = 0

    # 7. Rapor dosyalarını oku ve ilgili hücreleri doldur
    for idx, f in enumerate(sorted_files):
        r_akim = 3 + idx
        print(f"[*] Veriler işleniyor ({idx+1}/{total_files}): {os.path.basename(f)}")
        
        try:
            df = pd.read_excel(f)
        except Exception as e:
            print(f"[WARNING] {os.path.basename(f)} okunamadı: {e}")
            continue

        # Yaya, motosiklet, bisiklet ele
        df = df[~df['Tür'].str.lower().apply(lambda x: any(e in str(x).lower() for e in exclude_turs))]

        # Araçları say ve tabloya ekle
        for _, row in df.iterrows():
            entry_gate_val = get_gate_num(row['Giriş Kapısı'])
            exit_gate_val = get_gate_num(row['Çıkış Kapısı'])
            raw_tur = str(row['Tür']).lower().strip()

            mapped_tur = category_mapping.get(raw_tur)
            if mapped_tur and entry_gate_val and exit_gate_val:
                sheet_name = f"{entry_gate_val}. AKIM"
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Hedef sütunu col_mappings üzerinden dinamik bul
                    target_col = None
                    for c, (v_class, exit_num) in col_mappings_by_sheet[sheet_name].items():
                        if v_class == mapped_tur and exit_num == exit_gate_val:
                            target_col = c
                            break
                    if target_col:
                        curr_val = ws.cell(row=r_akim, column=target_col).value or 0
                        ws.cell(row=r_akim, column=target_col).value = curr_val + 1

    # 8. Tüm Akım sayfalarında satır toplamlarını hesapla
    for sheet_name in akim_sheets:
        ws = wb[sheet_name]
        col_map = col_mappings_by_sheet[sheet_name]
        tot_map = total_cols_by_sheet[sheet_name]
        
        for r in range(3, last_row_akim + 1):
            for mapped_tur, total_col_idx in tot_map.items():
                v_cols = [c for c, (v_class, _) in col_map.items() if v_class == mapped_tur]
                row_sum = sum(ws.cell(row=r, column=c).value or 0 for c in v_cols)
                ws.cell(row=r, column=total_col_idx).value = row_sum

    # Özet sayfasını bul (içinde "AKIM" geçmeyen tek sayfa özet sayfasıdır)
    ozet_sheet_name = None
    for name in wb.sheetnames:
        if "AKIM" not in name:
            ozet_sheet_name = name
            break
    if not ozet_sheet_name:
        ozet_sheet_name = wb.sheetnames[-1]
    ws_ozet = wb[ozet_sheet_name]
    # Sütun A birleştirilmiş alanlarını güncelle, verileri sıfırla ve doldur
    last_row_ozet = 1 + total_files
    total_row_ozet = last_row_ozet + 1
    
    # Dinamik satır ekleme/silme işlemi (ÖZET)
    if total_files < 30:
        copy_row_style(ws_ozet, 32, total_row_ozet)
        ws_ozet.delete_rows(total_row_ozet + 1, 32 - total_row_ozet)
    elif total_files > 30:
        ws_ozet.insert_rows(32, total_row_ozet - 32)
        for r in range(32, total_row_ozet):
            copy_row_style(ws_ozet, 31, r)
            
    # Sütun A birleştirilmiş alanlarını güncelle (ÖZET için satır 2'den başlar, last_row_ozet'e kadar)
    adjust_merged_column_a(ws_ozet, start_row=2, n_sabah=len(sabah_files), n_ogle=len(ogle_files), end_row=last_row_ozet)
    
    # Zaman ve verileri sıfırla/ilklendir
    for r in range(2, last_row_ozet + 1):
        for c in range(2, 10):
            ws_ozet.cell(row=r, column=c).value = 0

    # Her satırı doldur
    for r_ozet in range(2, last_row_ozet + 1):
        r_akim = r_ozet + 1
        
        # ZAMAN bilgisini ilk akım sayfasından kopyala
        ws_ozet.cell(row=r_ozet, column=2).value = wb[akim_sheets[0]].cell(row=r_akim, column=2).value
        
        # Verileri topla
        sums = {}
        for vehicle in category_mapping.values():
            v_sum = 0
            for sheet_name in akim_sheets:
                if sheet_name in wb.sheetnames:
                    tot_map = total_cols_by_sheet[sheet_name]
                    if vehicle in tot_map:
                        total_col_idx = tot_map[vehicle]
                        v_sum += wb[sheet_name].cell(row=r_akim, column=total_col_idx).value or 0
            sums[vehicle] = v_sum

        # ÖZET sayfasına yaz
        ws_ozet.cell(row=r_ozet, column=3).value = sums['AGIR_TASIT']
        ws_ozet.cell(row=r_ozet, column=4).value = sums['KAMYONET']
        ws_ozet.cell(row=r_ozet, column=5).value = sums['MINIBUS']
        ws_ozet.cell(row=r_ozet, column=6).value = sums['OTOBUS']
        ws_ozet.cell(row=r_ozet, column=7).value = sums['OTOMOBIL']
        ws_ozet.cell(row=r_ozet, column=8).value = sums['PANELVAN']
        ws_ozet.cell(row=r_ozet, column=9).value = sum(sums.values())

    # GÜN SONU TOPLAMLARI
    print(f"[*] GÜN SONU TOPLAMI hesaplanıyor (Satır {total_row_ozet})...")
    ws_ozet.cell(row=total_row_ozet, column=1).value = "GÜN SONU"
    ws_ozet.cell(row=total_row_ozet, column=2).value = "TOPLAMI"
    
    for c in range(3, 9):
        col_sum = sum(ws_ozet.cell(row=r, column=c).value or 0 for r in range(2, last_row_ozet + 1))
        ws_ozet.cell(row=total_row_ozet, column=c).value = col_sum
        
    ws_ozet.cell(row=total_row_ozet, column=9).value = sum(ws_ozet.cell(row=total_row_ozet, column=c).value or 0 for c in range(3, 9))

    # Grafiklerin Hücre Referanslarını Güncelle
    if len(ws_ozet._charts) >= 2:
        c0 = ws_ozet._charts[0] # Pie Chart
        c1 = ws_ozet._charts[1] # Line Chart
        
        # 1. Pasta Grafiği (Pie Chart) - Toplam Satırını referans alır
        c0.series[0].val.numRef.f = f"'{ws_ozet.title}'!$C${total_row_ozet}:$H${total_row_ozet}"
        c0.series[0].val.numCache = None
        
        # Grafik veri etiketlerini temizle (Seri adını gizleyip sadece kategori ve yüzdeyi göster)
        if c0.series[0].dLbls is not None:
            c0.series[0].dLbls.showSerName = False
            c0.series[0].dLbls.showVal = False
            c0.series[0].dLbls.showCatName = True
            c0.series[0].dLbls.showPercent = True
        
        # 2. Çizgi Grafiği (Line Chart) - Veri Satırlarını referans alır
        # Kategori formülü:
        for s in c1.series:
            s.cat.strRef.f = f"'{ws_ozet.title}'!$B$2:$B${last_row_ozet}"
            s.cat.strCache = None
            
        # Değer formülleri:
        cols_letters = ['C', 'D', 'E', 'F', 'G', 'H']
        for idx, s in enumerate(c1.series):
            if idx < len(cols_letters):
                col_letter = cols_letters[idx]
                s.val.numRef.f = f"'{ws_ozet.title}'!${col_letter}$2:${col_letter}${last_row_ozet}"
                s.val.numCache = None

    # Excel formüllerini etkin kıl
    wb.calculation.calcMode = 'auto'
    wb.calculation.forceFullCalc = True
    wb.calculation.fullCalcOnLoad = True

    # 10. Kaydet
    print(f"\n[INFO] Rapor kaydediliyor: {OUTPUT_PATH}")
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
        print(f"[BAŞARILI] Nihai rapor {OUTPUT_PATH} dosyasına başarıyla yazıldı!")
    except PermissionError:
        fallback_path = OUTPUT_PATH.replace(".xlsx", "_yeni.xlsx")
        print(f"[WARNING] '{OUTPUT_PATH}' açık veya kilitli olduğu için kaydedilemedi!")
        print(f"[INFO] Alternatif olarak şu dosyaya kaydediliyor: {fallback_path}")
        wb.save(fallback_path)
        print(f"[BAŞARILI] Nihai rapor {fallback_path} dosyasına başarıyla yazıldı!")

if __name__ == "__main__":
    main()
