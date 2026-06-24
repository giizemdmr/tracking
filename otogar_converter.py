import os
import glob
import openpyxl
import pandas as pd
from datetime import datetime, timedelta, time

TEMPLATE_PATH = "otogarkavsagi.xlsx"
OUTPUT_PATH = "ruspazari_genel.xlsx"
INPUT_DIR = "ruspazari"

# Base column offsets for vehicle types in Akım sheets
vehicle_bases = {
    'AGIR_TASIT': 0,   # start Col C (3)
    'KAMYONET': 4,     # start Col G (7)
    'MINIBUS': 8,      # start Col K (11)
    'OTOBUS': 12,      # start Col O (15)
    'OTOMOBIL': 16,    # start Col S (19)
    'PANELVAN': 20     # start Col W (23)
}

exclude_turs = ["bisiklet", "motosiklet", "yaya"]

def get_gate_num(gate_name):
    for char in str(gate_name):
        if char.isdigit():
            return int(char)
    return None

def parse_slot(time_str):
    try:
        parts = time_str.split('-')
        start_str = parts[0].strip()
        end_str = parts[1].strip()
        start_t = datetime.strptime(start_str, "%H:%M").time()
        end_t = datetime.strptime(end_str, "%H:%M").time()
        return start_t, end_t
    except Exception:
        return None, None

def main():
    print("="*60)
    print(" OTOGAR-STYLE CONVERTER: Rus Pazarı Veri Birleştirme")
    print("="*60)

    # 1. Raw Excel dosyalarını yükle ve birleştir
    files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    if not files:
        print(f"[ERROR] '{INPUT_DIR}' klasörü içinde excel dosyası bulunamadı!")
        return

    print(f"[INFO] '{INPUT_DIR}' klasöründen {len(files)} adet rapor okunuyor...")
    dfs = []
    for f in files:
        try:
            dfs.append(pd.read_excel(f))
        except Exception as e:
            print(f"[WARNING] {os.path.basename(f)} okunamadı: {e}")

    if not dfs:
        print("[ERROR] Hiçbir veri okunamadı!")
        return

    all_data = pd.concat(dfs, ignore_index=True)
    print(f"[INFO] Toplam {len(all_data)} satır veri birleştirildi.")

    # Yaya, motosiklet, bisiklet ele
    all_data = all_data[~all_data['Tür'].str.lower().apply(lambda x: any(e in str(x).lower() for e in exclude_turs))]
    print(f"[INFO] Yaya, motosiklet ve bisiklet filtrelemesinden sonra {len(all_data)} satır kaldı.")

    # Saatleri 3 saat 15 dakika geriye kaydır
    def shift_time(t_val):
        if pd.isna(t_val):
            return None
        try:
            t_parsed = datetime.strptime(str(t_val).strip(), "%H:%M:%S")
            return (t_parsed - timedelta(hours=3, minutes=15)).time()
        except Exception:
            try:
                t_parsed = datetime.strptime(str(t_val).strip(), "%H:%M")
                return (t_parsed - timedelta(hours=3, minutes=15)).time()
            except Exception:
                return None

    all_data['shifted_time'] = all_data['Zaman'].apply(shift_time)
    all_data = all_data.dropna(subset=['shifted_time'])
    print(f"[INFO] Zaman kaydırma ve filtreleme tamamlandı.")

    # 2. Şablon Excel Dosyasını Yükle
    if not os.path.exists(TEMPLATE_PATH):
        print(f"[ERROR] Şablon dosya bulunamadı: {TEMPLATE_PATH}")
        return

    print(f"[INFO] Şablon dosya yükleniyor: {TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    # 3. Sayfa İsimlerini Yeniden Adlandır (1, 2, 3 -> 2, 3, 4)
    wb['3. AKIM'].title = '4. AKIM_temp'
    wb['2. AKIM'].title = '3. AKIM_temp'
    wb['1. AKIM'].title = '2. AKIM_temp'

    wb['2. AKIM_temp'].title = '2. AKIM'
    wb['3. AKIM_temp'].title = '3. AKIM'
    wb['4. AKIM_temp'].title = '4. AKIM'
    print("[INFO] Sayfalar yeniden adlandırıldı: 2. AKIM, 3. AKIM, 4. AKIM")

    # 4. Sütun Başlıklarını Güncelle (Row 2 route isimleri)
    for entry_num in [2, 3, 4]:
        ws = wb[f"{entry_num}. AKIM"]
        # Update route headers in Row 2 (e.g. 2-2, 2-3, 2-4 for 2. AKIM)
        for base_offset in vehicle_bases.values():
            ws.cell(row=2, column=3 + base_offset).value = f"{entry_num}-2"
            ws.cell(row=2, column=4 + base_offset).value = f"{entry_num}-3"
            ws.cell(row=2, column=5 + base_offset).value = f"{entry_num}-4"
            ws.cell(row=2, column=6 + base_offset).value = "TOPLAM"

    # 5. Akım sayfalarını doldur
    akim_sheets = ['2. AKIM', '3. AKIM', '4. AKIM']
    for sheet_name in akim_sheets:
        ws = wb[sheet_name]
        entry_gate_val = get_gate_num(sheet_name)
        print(f"\n[*] Sayfa Dolduruluyor: {sheet_name} (Giriş Kapısı Gate_{entry_gate_val})")

        # C. Verileri temizle (Önceki şablondan kalan verileri sıfırla)
        for r in range(3, 35):
            for c in range(3, 27):
                ws.cell(row=r, column=c).value = 0

        # Filtrele (Giriş Kapısı == Gate_X)
        sheet_data = all_data[all_data['Giriş Kapısı'].apply(get_gate_num) == entry_gate_val]

        # Her satırı (zaman dilimini) işle
        for r in range(3, 35):
            time_str = ws.cell(row=r, column=2).value
            if not time_str:
                continue
            start_t, end_t = parse_slot(str(time_str))
            if start_t is None or end_t is None:
                continue

            # Zaman aralığına göre filtrele
            slot_data = sheet_data[(sheet_data['shifted_time'] >= start_t) & (sheet_data['shifted_time'] < end_t)]

            # Araçları say ve tabloya ekle
            for idx, row in slot_data.iterrows():
                raw_tur = str(row['Tür']).lower().strip()
                exit_gate_val = get_gate_num(row['Çıkış Kapısı'])

                # Tür eşleme
                if raw_tur == 'otomobil':
                    mapped_tur = 'OTOMOBIL'
                elif raw_tur == 'kamyonet':
                    mapped_tur = 'KAMYONET'
                elif raw_tur == 'minibus':
                    mapped_tur = 'MINIBUS'
                elif raw_tur == 'otobus':
                    mapped_tur = 'OTOBUS'
                elif raw_tur == 'panelvan':
                    mapped_tur = 'PANELVAN'
                elif raw_tur == 'agir_tasit':
                    mapped_tur = 'AGIR_TASIT'
                else:
                    mapped_tur = None

                if mapped_tur and exit_gate_val in [2, 3, 4]:
                    base_offset = vehicle_bases[mapped_tur]
                    # Exit gate 2 -> Col 3 + base_offset
                    # Exit gate 3 -> Col 4 + base_offset
                    # Exit gate 4 -> Col 5 + base_offset
                    col_idx = 3 + base_offset + (exit_gate_val - 2)
                    ws.cell(row=r, column=col_idx).value += 1

        # Satır Toplamlarını Hesapla (TOPLAM sütunları)
        for r in range(3, 35):
            for base_offset in vehicle_bases.values():
                r1 = ws.cell(row=r, column=3 + base_offset).value or 0
                r2 = ws.cell(row=r, column=4 + base_offset).value or 0
                r3 = ws.cell(row=r, column=5 + base_offset).value or 0
                ws.cell(row=r, column=6 + base_offset).value = r1 + r2 + r3

    # 6. ÖZET sayfasını doldur
    print("\n[*] ÖZET Sayfası Dolduruluyor...")
    ws_ozet = wb['ÖZET']
    for r_ozet in range(2, 34):
        r_akim = r_ozet + 1
        
        # Reset columns
        for c in range(3, 10):
            ws_ozet.cell(row=r_ozet, column=c).value = 0

        # Sum values across all three sheets
        sums = {}
        for vehicle, base_offset in vehicle_bases.items():
            # The total column for this vehicle in any akım sheet is Col 6 + base_offset
            col_total_idx = 6 + base_offset
            v_sum = 0
            for entry_num in [2, 3, 4]:
                v_sum += wb[f"{entry_num}. AKIM"].cell(row=r_akim, column=col_total_idx).value or 0
            sums[vehicle] = v_sum

        # Write to ÖZET sheet
        ws_ozet.cell(row=r_ozet, column=3).value = sums['AGIR_TASIT']
        ws_ozet.cell(row=r_ozet, column=4).value = sums['KAMYONET']
        ws_ozet.cell(row=r_ozet, column=5).value = sums['MINIBUS']
        ws_ozet.cell(row=r_ozet, column=6).value = sums['OTOBUS']
        ws_ozet.cell(row=r_ozet, column=7).value = sums['OTOMOBIL']
        ws_ozet.cell(row=r_ozet, column=8).value = sums['PANELVAN']
        
        # Genel Toplam
        ws_ozet.cell(row=r_ozet, column=9).value = sum(sums.values())

    # GÜN SONU TOPLAMI (Row 34) hesapla ve yaz
    print("[*] GÜN SONU TOPLAMI hesaplanıyor...")
    for c in range(3, 10):
        c_sum = sum(ws_ozet.cell(row=r, column=c).value or 0 for r in range(2, 34))
        ws_ozet.cell(row=34, column=c).value = c_sum

    # Excel formüllerinin otomatik hesaplanmasını ve grafiklerin güncellenmesini zorunlu kıl
    wb.calculation.calcMode = 'auto'
    wb.calculation.forceFullCalc = True
    wb.calculation.fullCalcOnLoad = True

    # 7. Kaydet
    print(f"\n[INFO] Genel rapor kaydediliyor: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print(f"[BAŞARILI] Genel rapor {OUTPUT_PATH} dosyasına yazıldı!")

if __name__ == "__main__":
    main()
